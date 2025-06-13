import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO

st.set_page_config(page_title="Excel Subtable Editor", layout="wide")
st.title("📊 Excel Subtable Editor")

# --- Configuration Constants ---
# Define the default auto-fill values for the toggle switch
AUTO_FILL_START_ROW = 23
AUTO_FILL_END_ROW = 36
AUTO_FILL_START_COL = 2
AUTO_FILL_END_COL = 5

# Define the order numbers to be automatically removed
ROWS_TO_AUTO_REMOVE = [8, 10, 13]

# Define the order numbers to be automatically combined and the new name
ROWS_TO_AUTO_COMBINE = [11, 12]
AUTO_COMBINED_ROW_NAME = "MT - Without FV"

# --- File Upload Section (Moved to Sidebar) ---
with st.sidebar:
    st.header("Upload Excel File")
    uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx", "xls"])
    st.markdown("---") # Separator in sidebar

# --- Utility Functions (Cached) ---
@st.cache_resource(ttl=3600)
def load_workbook_from_bytesio(file_buffer):
    """Loads an OpenPyXL workbook from a BytesIO object."""
    file_buffer.seek(0)
    return openpyxl.load_workbook(file_buffer, data_only=True)

@st.cache_data(ttl=3600)
def get_initial_dataframe(_workbook, sheet_name, start_row, end_row, start_col, end_col, use_first_row_as_header):
    """
    Extracts a DataFrame from a specified range within an OpenPyXL worksheet.
    Handles header logic and ensures unique column names.
    """
    ws = _workbook[sheet_name]

    # Ensure valid ranges, OpenPyXL is 1-indexed
    start_row = max(1, start_row)
    end_row = max(start_row, end_row) # Ensure end_row is not less than start_row
    start_col = max(1, start_col)
    end_col = max(start_col, end_col) # Ensure end_col is not less than start_col

    data = []
    try:
        # Iterate over cells and collect values
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col, values_only=True):
            data.append(list(row))
    except Exception as e:
        st.error(f"Error reading specified range from sheet: {e}. Please check your row/column inputs.")
        return pd.DataFrame() # Return empty DataFrame on error

    if not data:
        return pd.DataFrame() # Return empty if no data was read

    if use_first_row_as_header and len(data) > 0:
        raw_headers = list(data[0])
        rows = data[1:]
    else:
        # Generate default headers based on the number of columns in the selection
        raw_headers = [f"Column_{i+1}" for i in range(end_col - start_col + 1)]
        rows = data

    headers = []
    seen = {}
    for h in raw_headers:
        h_str = str(h) if h is not None and str(h).strip() != "" else "Unnamed"
        # Handle duplicate header names by appending a counter
        if h_str in seen:
            seen[h_str] += 1
            h_str = f"{h_str}_{seen[h_str]}"
        else:
            seen[h_str] = 0 # Initialize counter
        headers.append(h_str)
    
    # Ensure headers list matches the number of columns in rows
    # This can happen if some rows have fewer columns than the inferred max_col
    # Pad shorter rows with None if necessary to match header length
    adjusted_rows = []
    expected_cols = len(headers)
    for row in rows:
        if len(row) < expected_cols:
            adjusted_rows.append(list(row) + [None] * (expected_cols - len(row)))
        else:
            adjusted_rows.append(list(row[:expected_cols])) # Truncate if row has more columns than expected
            
    df_result = pd.DataFrame(adjusted_rows, columns=headers)
    df_result = df_result.dropna(how="all") # Drop rows where all values are NaN/empty

    # Add a default 'Order' column for reordering if it doesn't exist
    if 'Order' not in df_result.columns:
        df_result.insert(0, 'Order', range(1, len(df_result) + 1)) # Add at the beginning, 1-indexed

    return df_result

# --- Main App Logic ---
if uploaded_file is not None:
    try:
        with st.spinner("Loading Excel file..."):
            wb = load_workbook_from_bytesio(uploaded_file)
        
        sheet_names = wb.sheetnames
        st.success("File loaded successfully!")

        # --- Sheet Selection & Dimensions (Moved to Sidebar) ---
        with st.sidebar:
            st.header("Sheet Selection")
            selected_sheet = st.selectbox("Select a sheet", sheet_names, key="selected_sheet_sidebar")
            ws = wb[selected_sheet]
            max_row = ws.max_row
            max_column = ws.max_column
            st.info(f"Sheet dimensions: {max_row} rows, {max_column} columns")
            st.markdown("---") # Separator in sidebar

        # --- Subtable Selection Method ---
        st.markdown("### 🔍 Choose Subtable Selection Method")
        selection_method = st.radio(
            "How do you want to select the subtable?",
            ("Manual Range Input", "Auto-Detect by Blank Rows"),
            index=0,
            key="selection_method_radio"
        )

        df_initial = pd.DataFrame() # Initialize empty DataFrame

        # --- Initialize Session State for Manual Input Values ---
        # These will hold the current values of the number_inputs
        if "start_row_manual_val" not in st.session_state:
            st.session_state.start_row_manual_val = 1
        if "end_row_manual_val" not in st.session_state:
            st.session_state.end_row_manual_val = min(st.session_state.start_row_manual_val + 10, max_row)
        if "start_col_manual_val" not in st.session_state:
            st.session_state.start_col_manual_val = 1
        if "end_col_manual_val" not in st.session_state:
            st.session_state.end_col_manual_val = min(st.session_state.start_col_manual_val + 5, max_column)
        if "use_header_manual_val" not in st.session_state:
            st.session_state.use_header_manual_val = True

        # --- Auto-fill Toggle Switch ---
        # This toggle will update the session state values for manual inputs
        auto_fill_toggle = st.toggle(
            f"Auto-fill with predefined range (Rows {AUTO_FILL_START_ROW}-{AUTO_FILL_END_ROW}, Cols {AUTO_FILL_START_COL}-{AUTO_FILL_END_COL})",
            key="auto_fill_toggle_switch"
        )

        # If the toggle is active, update the session state values to the predefined ones
        if auto_fill_toggle:
            st.session_state.start_row_manual_val = AUTO_FILL_START_ROW
            st.session_state.end_row_manual_val = AUTO_FILL_END_ROW
            st.session_state.start_col_manual_val = AUTO_FILL_START_COL
            st.session_state.end_col_manual_val = AUTO_FILL_END_COL
            st.session_state.use_header_manual_val = True # Assume header for auto-filled data

        if selection_method == "Manual Range Input":
            st.markdown("#### Manual Subtable Range Selection")
            st.info("Enter the row and column numbers as they appear in Excel (1-indexed).")
            
            # Use session state values for the 'value' parameter of number_inputs
            start_row_manual = st.number_input(
                "Start Row", 
                min_value=1, 
                max_value=max_row, 
                value=st.session_state.start_row_manual_val, # Controlled by session state
                key="start_row_manual_input_key" # Unique key for the widget itself
            )
            end_row_manual = st.number_input(
                "End Row", 
                min_value=start_row_manual, # Dynamically adjust min_value
                max_value=max_row, 
                value=max(start_row_manual, st.session_state.end_row_manual_val), # Controlled by session state
                key="end_row_manual_input_key"
            )
            start_col_manual = st.number_input(
                "Start Column (A=1)", 
                min_value=1, 
                max_value=max_column, 
                value=st.session_state.start_col_manual_val, # Controlled by session state
                key="start_col_manual_input_key"
            )
            end_col_manual = st.number_input(
                "End Column", 
                min_value=start_col_manual, # Dynamically adjust min_value
                max_value=max_column, 
                value=max(start_col_manual, st.session_state.end_col_manual_val), # Controlled by session state
                key="end_col_manual_input_key"
            )
            use_first_row_as_header_manual = st.checkbox(
                "Use first row of selection as header", 
                value=st.session_state.use_header_manual_val, # Controlled by session state
                key="use_header_manual_input_key"
            )

            # Update session state with the current values from the number_inputs
            # This ensures manual edits persist across reruns
            st.session_state.start_row_manual_val = start_row_manual
            st.session_state.end_row_manual_val = end_row_manual
            st.session_state.start_col_manual_val = start_col_manual
            st.session_state.end_col_manual_val = end_col_manual
            st.session_state.use_header_manual_val = use_first_row_as_header_manual

            # Get the initial DataFrame based on the selected manual range
            df_initial = get_initial_dataframe(wb, selected_sheet, 
                                                start_row_manual, end_row_manual, 
                                                start_col_manual, end_col_manual, 
                                                use_first_row_as_header_manual)

        elif selection_method == "Auto-Detect by Blank Rows":
            st.markdown("#### Auto-Detecting Subtables")
            uploaded_file.seek(0) # Reset file pointer for pandas.read_excel
            # Read the entire sheet without header initially to detect blank rows
            full_df = pd.read_excel(uploaded_file, sheet_name=selected_sheet, header=None)

            non_empty_rows_indices = full_df.dropna(how='all').index.tolist()
            df_auto_detected = pd.DataFrame()

            if non_empty_rows_indices:
                # Find the first contiguous block of data
                first_data_row_idx_0based = non_empty_rows_indices[0]
                last_data_row_idx_0based = first_data_row_idx_0based
                for i in range(first_data_row_idx_0based + 1, len(full_df)):
                    if (i in non_empty_rows_indices) and ((i - 1) in non_empty_rows_indices):
                        last_data_row_idx_0based = i
                    else:
                        break
                
                # Extract the raw subtable based on detected rows
                sub_df_raw = full_df.iloc[first_data_row_idx_0based : last_data_row_idx_0based + 1].copy()
                # Detect non-empty columns within this subtable
                non_empty_cols_indices = sub_df_raw.dropna(axis=1, how='all').columns.tolist()

                if non_empty_cols_indices:
                    # Adjust to 1-based indexing for display
                    detected_start_row = first_data_row_idx_0based + 1
                    detected_end_row = last_data_row_idx_0based + 1
                    detected_start_col = non_empty_cols_indices[0] + 1
                    detected_end_col = non_empty_cols_indices[-1] + 1
                    
                    st.info(f"Auto-detected range: Rows {detected_start_row} to {detected_end_row}, Columns {detected_start_col} to {detected_end_col}")

                    # Option to confirm header for auto-detected table
                    use_auto_detected_header = st.checkbox("Use the first row of the auto-detected selection as header?", value=True, key="use_auto_header")

                    if use_auto_detected_header:
                        auto_headers = sub_df_raw.iloc[0, non_empty_cols_indices].tolist()
                        auto_rows = sub_df_raw.iloc[1:, non_empty_cols_indices].values.tolist()
                    else:
                        # If not using the first row as header, generate default headers
                        auto_headers = [f"Column_{i+1}" for i in range(len(non_empty_cols_indices))]
                        auto_rows = sub_df_raw.iloc[:, non_empty_cols_indices].values.tolist()

                    headers = []
                    seen = {}
                    for h in auto_headers:
                        h_str = str(h) if h is not None and str(h).strip() != "" else "Unnamed"
                        if h_str in seen:
                            seen[h_str] += 1
                            h_str = f"{h_str}_{seen[h_str]}"
                        else:
                            seen[h_str] = 0
                        headers.append(h_str)

                    # Ensure rows are correctly padded/truncated to match header length
                    adjusted_auto_rows = []
                    expected_cols_auto = len(headers)
                    for row in auto_rows:
                        if len(row) < expected_cols_auto:
                            adjusted_auto_rows.append(list(row) + [None] * (expected_cols_auto - len(row)))
                        else:
                            adjusted_auto_rows.append(list(row[:expected_cols_auto]))
                            
                    df_auto_detected = pd.DataFrame(adjusted_auto_rows, columns=headers)
                    df_auto_detected = df_auto_detected.dropna(how="all")

                    if 'Order' not in df_auto_detected.columns:
                        df_auto_detected.insert(0, 'Order', range(1, len(df_auto_detected) + 1))

                else:
                    st.warning("No contiguous data block found for auto-detection in columns within the detected rows.")
            else:
                st.warning("No non-empty rows found for auto-detection. The sheet might be entirely blank or formatted unusually.")

            df_initial = df_auto_detected

        # --- Session State Management for current_df and history ---
        # Create a unique ID to determine if the base data selection has changed
        current_data_selection_id = (
            f"{uploaded_file.file_id}-"
            f"{selected_sheet}-"
            f"{selection_method}-"
            f"{st.session_state.get('start_row_manual_val', '')}-"
            f"{st.session_state.get('end_row_manual_val', '')}-"
            f"{st.session_state.get('start_col_manual_val', '')}-"
            f"{st.session_state.get('end_col_manual_val', '')}-"
            f"{st.session_state.get('use_header_manual_val', '')}"
        )

        if "last_processed_file_id" not in st.session_state or st.session_state.last_processed_file_id != current_data_selection_id:
            st.session_state.current_df = df_initial.copy()
            st.session_state.history = []
            st.session_state.last_processed_file_id = current_data_selection_id
            st.info("New file, sheet, or selection parameters detected. Table and history reset.")
        elif st.session_state.current_df.empty and not df_initial.empty:
            # Re-initialize if the previous data was empty but new detection isn't
            st.session_state.current_df = df_initial.copy()
            st.session_state.history = []
            st.session_state.last_processed_file_id = current_data_selection_id
            st.info("Re-initializing table from file as previous data was empty.")

        # --- Auto-remove specific rows ---
        if not st.session_state.current_df.empty and 'Order' in st.session_state.current_df.columns:
            st.markdown("### 🗑️ Automatic Row Filtering")
            auto_remove_toggle = st.checkbox(
                f"Automatically remove rows with 'Order' numbers: {', '.join(map(str, ROWS_TO_AUTO_REMOVE))}",
                key="auto_remove_rows_toggle"
            )

            if auto_remove_toggle:
                original_row_count = len(st.session_state.current_df)
                
                # Convert 'Order' column to numeric, coercing errors to NaN, then fill NaN with a value that won't match
                # This ensures the comparison works even if 'Order' column has non-numeric entries
                df_temp = st.session_state.current_df.copy() # Operate on a copy
                df_temp['Order_numeric'] = pd.to_numeric(df_temp['Order'], errors='coerce')

                rows_to_keep_mask = ~df_temp['Order_numeric'].isin(ROWS_TO_AUTO_REMOVE)
                
                # Check if any rows are actually being removed before updating history
                if not rows_to_keep_mask.all(): # If not all rows are to be kept (i.e., some are to be removed)
                    st.session_state.history.append(st.session_state.current_df.copy()) # Save current state before removal
                    
                    # Apply filter and drop the temporary column from df_temp, then assign back
                    st.session_state.current_df = df_temp[rows_to_keep_mask].drop(columns=['Order_numeric']).reset_index(drop=True)
                    
                    removed_count = original_row_count - len(st.session_state.current_df)
                    st.success(f"Automatically removed {removed_count} row(s) based on predefined order numbers.")
                    st.rerun() # Rerun to display the filtered table
            st.markdown("---") # Separator for auto-filter options

        # --- Auto-combine specific rows ---
        # Ensure we have enough rows and the 'Order' column before attempting to combine
        if not st.session_state.current_df.empty and 'Order' in st.session_state.current_df.columns and len(ROWS_TO_AUTO_COMBINE) > 1:
            st.markdown("### 🔗 Automatic Row Combination")
            auto_combine_toggle = st.checkbox(
                f"Automatically combine rows with 'Order' numbers: {', '.join(map(str, ROWS_TO_AUTO_COMBINE))} and rename to '{AUTO_COMBINED_ROW_NAME}'",
                key="auto_combine_rows_toggle"
            )

            if auto_combine_toggle:
                # Convert 'Order' column to numeric for robust comparison
                df_temp_combine = st.session_state.current_df.copy() # Operate on a copy
                df_temp_combine['Order_numeric_combine'] = pd.to_numeric(df_temp_combine['Order'], errors='coerce')

                # Find the actual indices of rows to combine based on 'Order_numeric_combine'
                # Ensure we only pick rows that are currently present and match
                indices_to_combine = df_temp_combine[df_temp_combine['Order_numeric_combine'].isin(ROWS_TO_AUTO_COMBINE)].index.tolist()

                if len(indices_to_combine) >= 2: # Only combine if at least two target rows exist
                    st.session_state.history.append(st.session_state.current_df.copy()) # Save current state

                    combined_row_data = {}
                    selected_df_for_auto_combine = st.session_state.current_df.loc[indices_to_combine]

                    for col_idx, col in enumerate(st.session_state.current_df.columns):
                        if pd.api.types.is_numeric_dtype(st.session_state.current_df[col]):
                            combined_row_data[col] = selected_df_for_auto_combine[col].sum()
                        else:
                            # Join non-numeric values, handling NaNs
                            joined_value = " / ".join(selected_df_for_auto_combine[col].dropna().astype(str).tolist())
                            combined_row_data[col] = joined_value

                    # Crucial: Ensure the first column gets the desired name if it exists and is not 'Order'
                    # The 'Order' column should ideally retain its numeric property or be handled explicitly.
                    # We assume the user wants the new name in the *first non-Order* column or the first column if Order doesn't exist.
                    if st.session_state.current_df.columns[0] != 'Order':
                        combined_row_data[st.session_state.current_df.columns[0]] = AUTO_COMBINED_ROW_NAME
                    elif len(st.session_state.current_df.columns) > 1: # If 'Order' is the first, try the second
                         # Find the first column that's not 'Order' if 'Order' is the first column.
                        first_non_order_col = next((col for col in st.session_state.current_df.columns if col != 'Order'), None)
                        if first_non_order_col:
                            combined_row_data[first_non_order_col] = AUTO_COMBINED_ROW_NAME


                    # Create a new DataFrame for the single combined row
                    combined_df_new = pd.DataFrame([combined_row_data], columns=st.session_state.current_df.columns)
                    
                    # Remove the original selected rows and add the new combined row
                    remaining_df = st.session_state.current_df.drop(index=indices_to_combine).reset_index(drop=True)
                    st.session_state.current_df = pd.concat([remaining_df, combined_df_new], ignore_index=True)
                    
                    st.success(f"Automatically combined rows with Order {', '.join(map(str, ROWS_TO_AUTO_COMBINE))} into '{AUTO_COMBINED_ROW_NAME}'.")
                    st.rerun()
                else:
                    st.warning(f"Could not auto-combine. Found {len(indices_to_combine)} row(s) with order numbers {', '.join(map(str, ROWS_TO_AUTO_COMBINE))}. At least 2 are required.")
            st.markdown("---") # Separator for auto-filter options


        # --- Display and Editing UI ---
        if not st.session_state.current_df.empty:
            st.subheader("✏️ Edit Table and Reorder Rows")
            st.info("To reorder rows, edit the numbers in the 'Order' column. To delete a row, click the 'X' button on the right of the row in the table.")

            # Ensure 'Order' column is numeric for proper sorting
            st.session_state.current_df['Order'] = pd.to_numeric(st.session_state.current_df['Order'], errors='coerce').fillna(0).astype(int)

            edited_df = st.data_editor(
                st.session_state.current_df,
                num_rows="dynamic", # Allows adding/deleting rows directly in the editor
                use_container_width=True,
                column_config={
                    "Order": st.column_config.NumberColumn(
                        "Order",
                        help="Assign a number to reorder rows.",
                        default=0, # Default value for new rows' order
                        step=1,
                        format="%d"
                    )
                },
                key="main_data_editor" # Unique key for the data editor
            )

            # Check if edited_df is different from current_df
            # This triggers a history save and success message
            if not edited_df.equals(st.session_state.current_df):
                st.session_state.history.append(st.session_state.current_df.copy())
                st.session_state.current_df = edited_df.copy()
                st.success("Changes detected. Apply order or continue editing.")
                st.rerun() # Rerun to reflect changes immediately and prevent edit conflicts

            if st.button("Apply New Row Order"):
                if 'Order' in st.session_state.current_df.columns:
                    temp_df = st.session_state.current_df.copy()

                    # Handle duplicate order numbers by adding a decimal for stable sort
                    # For example, if two rows have order 5, they become 5.0 and 5.1
                    temp_df['Order_temp_sort'] = temp_df['Order']
                    if temp_df['Order_temp_sort'].duplicated().any():
                        # Add a unique identifier for duplicates to maintain relative order
                        # This creates values like 5.0, 5.1, 5.2 if there are multiple 5s
                        temp_df['Order_temp_sort'] = temp_df['Order'].astype(str) + '.' + temp_df.groupby('Order_temp_sort').cumcount().astype(str)
                        temp_df['Order_temp_sort'] = pd.to_numeric(temp_df['Order_temp_sort'], errors='coerce')

                    st.session_state.current_df = temp_df.sort_values(by='Order_temp_sort', ascending=True).drop(columns=['Order_temp_sort']).reset_index(drop=True)
                    st.success("Rows reordered successfully!")
                    st.rerun() # Rerun to update the displayed dataframe with new order

                else:
                    st.warning("No 'Order' column found to reorder rows.")

            st.subheader("🔗 Combine Rows")
            st.write("Current table row indices:")
            # Display current indices for user reference
            st.dataframe(st.session_state.current_df.index.to_frame(name='Index'), use_container_width=True)
            st.info("Please select rows using the indices displayed above for the *current table*.")

            selected_rows_to_combine = st.multiselect(
                "Select rows to combine (by current table index)",
                st.session_state.current_df.index.tolist(),
                key="combine_rows_multiselect"
            )
            custom_name_for_combined_row = st.text_input("Custom name for the new combined row", value="Combined Row", key="custom_combined_row_name")

            if st.button("Combine Selected Rows"):
                if selected_rows_to_combine:
                    st.session_state.history.append(st.session_state.current_df.copy()) # Save current state

                    combined_row_data = {}
                    selected_df_for_combine = st.session_state.current_df.loc[selected_rows_to_combine]

                    for col in st.session_state.current_df.columns:
                        if pd.api.types.is_numeric_dtype(st.session_state.current_df[col]):
                            combined_row_data[col] = selected_df_for_combine[col].sum()
                        else:
                            # Join non-numeric values, handling NaNs
                            combined_row_data[col] = " / ".join(selected_df_for_combine[col].dropna().astype(str).tolist())
                            # Ensure the first column (often descriptive) gets the custom name
                            if col == st.session_state.current_df.columns[0]:
                                combined_row_data[col] = custom_name_for_combined_row

                    # Create a new DataFrame for the single combined row
                    combined_df = pd.DataFrame([combined_row_data], columns=st.session_state.current_df.columns)
                    
                    # Remove the original selected rows and add the new combined row
                    remaining_df = st.session_state.current_df.drop(index=selected_rows_to_combine).reset_index(drop=True)
                    st.session_state.current_df = pd.concat([remaining_df, combined_df], ignore_index=True)
                    st.success("Rows combined successfully.")
                    st.rerun()

                else:
                    st.warning("No rows selected to combine.")

            st.subheader("🧬 Merge Columns")
            selected_cols_to_merge = st.multiselect("Select columns to merge", st.session_state.current_df.columns.tolist(), key="merge_cols_multiselect")
            new_merged_col_name = st.text_input("New column name for merged data", value="MergedColumn", key="new_merged_col_name_input")
            
            if st.button("Merge Selected Columns"):
                if selected_cols_to_merge and len(selected_cols_to_merge) >= 2:
                    # Check if the new column name conflicts with existing non-selected columns
                    if new_merged_col_name in st.session_state.current_df.columns and new_merged_col_name not in selected_cols_to_merge:
                        st.error(f"Column '{new_merged_col_name}' already exists. Please choose a different name or include it in columns to merge if you intend to overwrite.")
                    else:
                        st.session_state.history.append(st.session_state.current_df.copy()) # Save current state
                        
                        # Create the new merged column by joining string representations
                        # Handle NaNs: dropna() removes NaNs before joining
                        st.session_state.current_df[new_merged_col_name] = (
                            st.session_state.current_df[selected_cols_to_merge]
                            .astype(str)
                            .agg(lambda x: " / ".join(x.dropna()), axis=1) # Join only non-NaN strings
                        )
                        # Drop the original columns that were merged
                        st.session_state.current_df.drop(columns=selected_cols_to_merge, inplace=True)
                        st.success(f"Columns merged into '{new_merged_col_name}'")
                        st.rerun()
                else:
                    st.warning("Please select at least two columns to merge.")

            if st.button("Undo Last Action"):
                if st.session_state.history:
                    st.session_state.current_df = st.session_state.history.pop()
                    st.success("Undo successful. Table restored to previous state.")
                    st.rerun() # Rerun to update the displayed dataframe
                else:
                    st.warning("No previous state to undo. History is empty.")

            st.subheader("📋 Final Edited Table")
            # Display the final, non-all-NA rows of the table
            final_df = st.session_state.current_df.dropna(how="all").reset_index(drop=True)
            st.dataframe(final_df, use_container_width=True)

            st.subheader("📥 Download Modified Table")
            def to_excel(df_to_save):
                """Converts a DataFrame to an Excel file in BytesIO object."""
                output = BytesIO()
                if not df_to_save.empty:
                    with pd.ExcelWriter(output, engine="openpyxl") as writer:
                        df_to_save.to_excel(writer, index=False, sheet_name="ModifiedTable")
                output.seek(0)
                return output

            # Generate Excel data for download
            excel_data = to_excel(final_df)
            st.download_button(
                "Download as Excel",
                data=excel_data,
                file_name="modified_subtable.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        else:
            st.info("No data found for the selected range/auto-detection. Please adjust your selection or upload a different file.")

    except Exception as e:
        st.error(f"An unexpected error occurred while processing the Excel file: {e}")
        st.exception(e) # Display full traceback for debugging
        st.info("Please ensure it's a valid Excel file with readable content and try again.")
else:
    st.info("Please upload an Excel file to begin.")
